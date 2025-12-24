import time
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# --- CONFIGURACIÓN ---
RUTA_EXCEL = r"C:\Users\Giral\OneDrive\Documentos\CATQALOGO MOTOS\catalogo_completo\CATALOGO TEMPLO GARAGE.xlsm"
NOMBRE_HOJA = "MUNDIMOTOS_COMPLETO_20251206_14"
COL_URL = 6           # Columna F (product_url)
COL_DESC = 15         # Columna O (descripcion)
FILA_INICIO = 2       # Primera fila con datos
DELAY = 1             # Segundos de espera entre peticiones (para no saturar el servidor)
# --------------------

def obtener_descripcion(url):
    """
    Obtiene la descripción de un producto desde la metaetiqueta 'description' de la página web.
    """
    try:
        # 1. Realizar la petición HTTP
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }
        resp = requests.get(url, headers=headers, timeout=10)
        resp.raise_for_status()  # Lanza excepción si hay error HTTP

        # 2. Parsear el HTML
        soup = BeautifulSoup(resp.text, 'html.parser')

        # 3. Buscar la metaetiqueta 'description'
        meta_desc = soup.find('meta', attrs={'name': 'description'})
        if meta_desc and meta_desc.get('content'):
            return meta_desc['content'].strip()

        # Si no se encuentra la metaetiqueta, retornar una cadena vacía
        return ""

    except requests.exceptions.RequestException as e:
        print(f"  Error al obtener {url}: {e}")
        return ""
    except Exception as e:
        print(f"  Error inesperado en {url}: {e}")
        return ""

def main():
    # 1. Cargar el libro de Excel
    print(f"Cargando archivo: {RUTA_EXCEL}")
    wb = load_workbook(RUTA_EXCEL, keep_vba=True)  # keep_vba=True para mantener macros
    ws = wb[NOMBRE_HOJA]

    # 2. Determinar la última fila con datos en la columna de URLs
    ultima_fila = ws.max_row
    print(f"Total de filas a procesar: {ultima_fila - FILA_INICIO + 1}")

    # 3. Recorrer cada fila
    for fila in range(FILA_INICIO, ultima_fila + 1):
        url = ws.cell(row=fila, column=COL_URL).value

        # Si la celda de URL está vacía, saltar
        if not url or not isinstance(url, str):
            print(f"Fila {fila}: URL vacía o no válida. Saltando.")
            ws.cell(row=fila, column=COL_DESC, value="")
            continue

        print(f"Fila {fila}: Procesando {url[:60]}...")
        descripcion = obtener_descripcion(url)

        # 4. Escribir la descripción en la columna O
        ws.cell(row=fila, column=COL_DESC, value=descripcion)

        # 5. Esperar un tiempo para no saturar el servidor
        time.sleep(DELAY)

    # 6. Guardar los cambios en el mismo archivo
    wb.save(RUTA_EXCEL)
    print("¡Proceso completado! Las descripciones se han guardado en el archivo.")

if __name__ == "__main__":
    main()